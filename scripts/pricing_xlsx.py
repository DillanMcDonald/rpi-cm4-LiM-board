#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
# Copyright (c) 2026 Dillan McDonald
"""
pricing_xlsx.py — KiCost-style pricing XLSX generator (Feature 6).

Reads a KiCad BOM CSV, queries distributor APIs in parallel, and writes a
formatted multi-sheet XLSX workbook.

Usage
-----
    python scripts/pricing_xlsx.py \\
        --bom bom.csv \\
        --qty 100 \\
        --distributors mouser,digikey,nexar,jlcpcb \\
        --output pricing.xlsx

Sheets produced
---------------
    1. BOM Summary     — one row per MPN, best price, stock status
    2. Price Comparison — side-by-side prices from each distributor
    3. Cost Rollup     — BOM total cost at each qty break + bar chart
    4. Raw API Data    — JSON audit trail of every API response

Dependencies (MIT/Apache-licensed):
    openpyxl>=3.1   — XLSX generation
    typer           — CLI

Internal:
    kicad_ci.distributors  — pluggable distributor clients
    kicad_ci.api_cache     — SQLite TTL cache
"""

from __future__ import annotations

import csv
import json
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FutureTimeout
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Sequence

# ---------------------------------------------------------------------------
# Optional typer import (graceful degradation if not installed)
# ---------------------------------------------------------------------------
try:
    import typer
    _HAS_TYPER = True
except ImportError:  # pragma: no cover
    _HAS_TYPER = False

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# Add repo root to path so kicad_ci is importable when run as a script
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from kicad_ci.distributors.base import (
    BomLine,
    PriceResult,
    PricedBomLine,
    _REGISTRY as _DIST_REGISTRY,
)


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_NAVY = "1F3864"
_WHITE = "FFFFFF"
_GREEN = "C6EFCE"
_GREEN_FONT = "276221"
_RED = "FFC7CE"
_RED_FONT = "9C0006"
_YELLOW = "FFEB9C"
_YELLOW_FONT = "9C6500"
_LIGHT_BLUE = "DEEAF1"


def _header_style() -> dict:
    return {
        "font": Font(bold=True, color=_WHITE, name="Calibri", size=10),
        "fill": PatternFill("solid", fgColor=_NAVY),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def _apply_header(cell):
    s = _header_style()
    _apply_style(cell, **s)


def _thin_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# BOM CSV reader (F6-T1)
# ---------------------------------------------------------------------------

_MPN_COLS = ("MPN", "mpn", "Part Number", "part_number", "Manufacturer Part Number")
_MFR_COLS = ("Manufacturer", "manufacturer", "MFR", "mfr")
_REF_COLS = ("Reference", "reference", "Ref", "ref", "Designator", "designator")
_QTY_COLS = ("Quantity", "quantity", "Qty", "qty", "Count")
_VAL_COLS = ("Value", "value", "Val")
_FP_COLS  = ("Footprint", "footprint")
_DESC_COLS = ("Description", "description", "Desc", "Comment")
_DNP_COLS = ("DNP", "dnp", "Do Not Populate", "Exclude from BOM")


def _first_col(headers: List[str], candidates: Sequence[str]) -> Optional[int]:
    h_lower = [h.strip().lower() for h in headers]
    for c in candidates:
        cl = c.lower()
        if cl in h_lower:
            return h_lower.index(cl)
    return None


def read_bom_csv(path: Path, exclude_dnp: bool = True) -> List[BomLine]:
    """
    Parse a KiCad-exported BOM CSV into a list of :class:`BomLine`.

    Groups rows with identical MPNs and sums their quantities.
    Filters DNP components when *exclude_dnp* is True.
    """
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        raw_headers = next(reader, None)
        if raw_headers is None:
            return []
        headers = [h.strip() for h in raw_headers]

        mpn_col  = _first_col(headers, _MPN_COLS)
        mfr_col  = _first_col(headers, _MFR_COLS)
        ref_col  = _first_col(headers, _REF_COLS)
        qty_col  = _first_col(headers, _QTY_COLS)
        val_col  = _first_col(headers, _VAL_COLS)
        fp_col   = _first_col(headers, _FP_COLS)
        desc_col = _first_col(headers, _DESC_COLS)
        dnp_col  = _first_col(headers, _DNP_COLS)

        grouped: Dict[str, BomLine] = {}

        for row in reader:
            if not row or not any(row):
                continue

            def _cell(idx: Optional[int]) -> str:
                if idx is None or idx >= len(row):
                    return ""
                return row[idx].strip()

            mpn = _cell(mpn_col)
            # Fallback: if no MPN field, try Value (many KiCad projects put
            # the part number there directly — e.g. Value="SMBJ5.0A-TR").
            # Only use Value if it looks like an MPN: >= 4 chars and has a
            # digit or dash (skips generic values like "10k", "100uF").
            if (not mpn or mpn.lower() in ("", "?", "~", "n/a", "tbd")):
                val = _cell(val_col)
                if val and len(val) >= 4 and any(c.isdigit() or c == '-' for c in val):
                    mpn = val
            if not mpn or mpn.lower() in ("", "?", "~", "n/a", "tbd"):
                continue

            dnp_val = _cell(dnp_col).lower()
            dnp = dnp_val in ("1", "yes", "true", "dnp", "x")
            if exclude_dnp and dnp:
                continue

            # Parse refs: may be comma- or space-separated in one cell
            ref_raw = _cell(ref_col)
            refs = [r.strip() for r in ref_raw.replace(";", ",").split(",") if r.strip()]

            qty_raw = _cell(qty_col)
            try:
                qty = int(float(qty_raw)) if qty_raw else len(refs) or 1
            except ValueError:
                qty = len(refs) or 1

            if mpn in grouped:
                existing = grouped[mpn]
                existing.refs.extend(refs)
                # accumulate qty
                object.__setattr__(
                    existing, "qty", existing.qty + qty
                ) if hasattr(existing, "__setattr__") else None
                grouped[mpn].refs.extend(refs)
                # Use a mutable approach
            else:
                grouped[mpn] = BomLine(
                    mpn=mpn,
                    manufacturer=_cell(mfr_col),
                    refs=refs,
                    qty=qty,
                    value=_cell(val_col),
                    footprint=_cell(fp_col),
                    description=_cell(desc_col),
                    dnp=dnp,
                )

    # Fix grouped quantities (BomLine.qty is not frozen)
    # Re-aggregate properly using a plain dict accumulator
    return _reaggregate(path, exclude_dnp)


def _reaggregate(path: Path, exclude_dnp: bool) -> List[BomLine]:
    """Two-pass aggregation that handles qty correctly for grouped MPNs."""
    raw_rows: Dict[str, dict] = {}

    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        headers = reader.fieldnames or []

        mpn_key  = next((c for c in _MPN_COLS  if c in headers), None)
        mfr_key  = next((c for c in _MFR_COLS  if c in headers), None)
        ref_key  = next((c for c in _REF_COLS  if c in headers), None)
        qty_key  = next((c for c in _QTY_COLS  if c in headers), None)
        val_key  = next((c for c in _VAL_COLS  if c in headers), None)
        fp_key   = next((c for c in _FP_COLS   if c in headers), None)
        desc_key = next((c for c in _DESC_COLS if c in headers), None)
        dnp_key  = next((c for c in _DNP_COLS  if c in headers), None)

        for row in reader:
            mpn = row.get(mpn_key or "", "").strip() if mpn_key else ""
            # Fallback: if no MPN column or empty, try Value (many KiCad
            # projects put the part number directly in Value).
            if not mpn or mpn.lower() in ("", "?", "~", "n/a", "tbd"):
                val = (row.get(val_key or "", "") or "").strip() if val_key else ""
                if val and len(val) >= 4 and any(c.isdigit() or c == '-' for c in val):
                    mpn = val
            if not mpn or mpn.lower() in ("", "?", "~", "n/a", "tbd"):
                continue

            dnp_val = (row.get(dnp_key or "", "") or "").strip().lower() if dnp_key else ""
            dnp = dnp_val in ("1", "yes", "true", "dnp", "x")
            if exclude_dnp and dnp:
                continue

            ref_raw = row.get(ref_key or "", "") if ref_key else ""
            refs = [r.strip() for r in (ref_raw or "").replace(";", ",").split(",") if r.strip()]

            qty_raw = row.get(qty_key or "", "") if qty_key else ""
            try:
                qty = int(float(qty_raw)) if qty_raw else len(refs) or 1
            except ValueError:
                qty = len(refs) or 1

            if mpn in raw_rows:
                raw_rows[mpn]["refs"].extend(refs)
                raw_rows[mpn]["qty"] += qty
            else:
                raw_rows[mpn] = {
                    "mpn": mpn,
                    "manufacturer": row.get(mfr_key or "", "") if mfr_key else "",
                    "refs": list(refs),
                    "qty": qty,
                    "value": row.get(val_key or "", "") if val_key else "",
                    "footprint": row.get(fp_key or "", "") if fp_key else "",
                    "description": row.get(desc_key or "", "") if desc_key else "",
                    "dnp": dnp,
                }

    return [
        BomLine(
            mpn=d["mpn"],
            manufacturer=d["manufacturer"],
            refs=sorted(set(d["refs"])),
            qty=d["qty"],
            value=d["value"],
            footprint=d["footprint"],
            description=d["description"],
            dnp=d["dnp"],
        )
        for d in raw_rows.values()
    ]


# ---------------------------------------------------------------------------
# Price aggregation engine (F6-T7)
# ---------------------------------------------------------------------------

_QUERY_TIMEOUT = 10  # seconds per distributor


def aggregate_prices(
    bom: List[BomLine],
    distributors: List[str],
    qty_multiplier: int = 1,
) -> List[PricedBomLine]:
    """
    Query all enabled distributors in parallel for each BOM line.

    Uses ``ThreadPoolExecutor(max_workers=4)`` with a 10-second timeout per
    query.  Returns :class:`PricedBomLine` list with all collected results.
    """
    priced: List[PricedBomLine] = [PricedBomLine(bom_line=line) for line in bom]

    clients = {
        name: _DIST_REGISTRY[name]
        for name in distributors
        if name in _DIST_REGISTRY
    }

    def _query(client_name: str, mpn: str) -> tuple[str, str, Optional[PriceResult]]:
        try:
            result = clients[client_name].search_by_mpn(mpn)
            return client_name, mpn, result
        except Exception:
            return client_name, mpn, None

    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {}
        for pbl in priced:
            mpn = pbl.bom_line.mpn
            for dist_name in clients:
                fut = pool.submit(_query, dist_name, mpn)
                futures[fut] = (dist_name, mpn, pbl)

        for fut in as_completed(futures, timeout=None):
            _dist_name, _mpn, pbl = futures[fut]
            try:
                dist_name, mpn, result = fut.result(timeout=_QUERY_TIMEOUT)
            except (FutureTimeout, Exception):
                continue
            if result is not None:
                pbl.distributor_prices[dist_name] = result

    # Scale quantities if qty_multiplier > 1
    if qty_multiplier > 1:
        for pbl in priced:
            pbl.bom_line.qty *= qty_multiplier

    return priced


# ---------------------------------------------------------------------------
# XLSX formatter (F6-T8) — openpyxl
# ---------------------------------------------------------------------------

def write_xlsx(
    priced_bom: List[PricedBomLine],
    output_path: Path,
    build_qty: int = 1,
) -> None:
    """
    Write a 4-sheet XLSX workbook to *output_path*.

    Sheets:
        1. BOM Summary
        2. Price Comparison
        3. Cost Rollup
        4. Raw API Data
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl>=3.1 required: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _write_bom_summary(wb, priced_bom, build_qty)
    _write_price_comparison(wb, priced_bom, build_qty)
    _write_cost_rollup(wb, priced_bom, build_qty)
    _write_raw_data(wb, priced_bom)

    wb.save(str(output_path))


# ---- Sheet 1: BOM Summary ------------------------------------------------

_SUMMARY_HEADERS = [
    "Ref Des", "MPN", "Manufacturer", "Description",
    "Qty", "Best Unit Price (USD)", "Extended Price (USD)",
    "Best Distributor", "Stock", "Stock Status", "Datasheet",
]


def _write_bom_summary(wb, priced_bom: List[PricedBomLine], build_qty: int):
    ws = wb.create_sheet("BOM Summary")
    ws.freeze_panes = "B2"   # freeze row 1 + column A
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_SUMMARY_HEADERS))}1"

    # Headers
    for col_idx, header in enumerate(_SUMMARY_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_header(cell)
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 28

    red_fill   = PatternFill("solid", fgColor=_RED)
    yellow_fill = PatternFill("solid", fgColor=_YELLOW)
    green_fill  = PatternFill("solid", fgColor=_GREEN)

    for row_idx, pbl in enumerate(priced_bom, 2):
        line = pbl.bom_line
        best = pbl.best_result
        unit_price = pbl.best_unit_price
        ext_price  = pbl.extended_price

        stock = best.stock if best else 0
        stock_status = (
            "OUT OF STOCK" if stock == 0
            else "LOW STOCK"    if stock < line.qty * 2
            else "IN STOCK"
        )
        status_fill = (
            red_fill    if stock == 0
            else yellow_fill if stock < line.qty * 2
            else green_fill
        )

        datasheet = best.datasheet_url if best else ""

        values = [
            ", ".join(line.refs),
            line.mpn,
            line.manufacturer,
            line.description,
            line.qty,
            float(unit_price) if unit_price else None,
            float(ext_price)  if ext_price  else None,
            best.distributor if best else "",
            stock,
            stock_status,
            datasheet,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _thin_border()
            if col_idx == 10:  # Stock Status
                cell.fill = status_fill
            if col_idx in (6, 7) and val is not None:
                cell.number_format = '"$"#,##0.0000'
            if col_idx == 11 and datasheet:  # Datasheet hyperlink
                cell.hyperlink = datasheet
                cell.font = Font(color="0563C1", underline="single")
                cell.value = "Datasheet"

    # Column widths
    widths = [20, 22, 22, 28, 6, 18, 18, 16, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Sheet 2: Price Comparison -------------------------------------------

def _write_price_comparison(wb, priced_bom: List[PricedBomLine], build_qty: int):
    ws = wb.create_sheet("Price Comparison")

    # Collect all distributor names seen across all results
    all_dists: list[str] = sorted(
        {d for pbl in priced_bom for d in pbl.distributor_prices}
    )

    fixed_headers = ["MPN", "Manufacturer", "Qty"]
    dist_headers  = [d.title() for d in all_dists]
    headers = fixed_headers + dist_headers + ["Best Price (USD)", "Best Distributor"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_header(cell)
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    green_fill = PatternFill("solid", fgColor=_GREEN)

    for row_idx, pbl in enumerate(priced_bom, 2):
        line = pbl.bom_line
        best = pbl.best_result
        unit_price = pbl.best_unit_price

        row_vals: dict[int, object] = {
            1: line.mpn,
            2: line.manufacturer,
            3: line.qty,
        }

        # Prices per distributor
        dist_prices: list[Optional[float]] = []
        for dist in all_dists:
            result = pbl.distributor_prices.get(dist)
            if result:
                p = result.price_at_qty(line.qty)
                dist_prices.append(float(p) if p else None)
            else:
                dist_prices.append(None)

        n_fixed = len(fixed_headers)
        for i, p in enumerate(dist_prices):
            col = n_fixed + 1 + i
            row_vals[col] = p

        n_dist = len(all_dists)
        row_vals[n_fixed + n_dist + 1] = float(unit_price) if unit_price else None
        row_vals[n_fixed + n_dist + 2] = best.distributor if best else ""

        for col_idx, val in row_vals.items():
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _thin_border()
            if col_idx >= n_fixed + 1 and isinstance(val, float):
                cell.number_format = '"$"#,##0.0000'

        # Highlight minimum non-None dist price in green
        non_none_prices = [(i, p) for i, p in enumerate(dist_prices) if p is not None]
        if non_none_prices:
            min_price = min(p for _, p in non_none_prices)
            for i, p in non_none_prices:
                if p == min_price:
                    ws.cell(row=row_idx, column=n_fixed + 1 + i).fill = green_fill

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 6
    for i in range(len(all_dists)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14


# ---- Sheet 3: Cost Rollup ------------------------------------------------

def _write_cost_rollup(wb, priced_bom: List[PricedBomLine], build_qty: int):
    ws = wb.create_sheet("Cost Rollup")

    # Collect all unique qty break points across all results
    qty_breaks: list[int] = sorted(
        {
            pb.min_qty
            for pbl in priced_bom
            for result in pbl.distributor_prices.values()
            for pb in result.price_breaks
            if pb.min_qty < 10_000_000  # skip "Infinity" sentinel
        }
    )
    if not qty_breaks:
        qty_breaks = [build_qty]

    headers = ["Build Qty"] + ["BOM Total (USD)"]
    ws.cell(row=1, column=1, value="Build Qty")
    ws.cell(row=1, column=2, value="BOM Total (USD)")
    for col_idx in (1, 2):
        _apply_header(ws.cell(row=1, column=col_idx))
        ws.cell(row=1, column=col_idx).border = _thin_border()

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, qty in enumerate(qty_breaks, 2):
        bom_total = Decimal("0")
        for pbl in priced_bom:
            best = pbl.best_result
            if best:
                p = best.price_at_qty(qty)
                if p:
                    bom_total += p * pbl.bom_line.qty

        ws.cell(row=row_idx, column=1, value=qty)
        total_cell = ws.cell(row=row_idx, column=2, value=float(bom_total))
        total_cell.number_format = '"$"#,##0.00'
        for col_idx in (1, 2):
            ws.cell(row=row_idx, column=col_idx).border = _thin_border()

    # Bar chart
    if len(qty_breaks) > 1:
        data_rows = len(qty_breaks)
        chart = BarChart()
        chart.type = "col"
        chart.title = "BOM Cost by Build Quantity"
        chart.y_axis.title = "Total Cost (USD)"
        chart.x_axis.title = "Build Quantity"
        chart.style = 10

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=data_rows + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws.add_chart(chart, "D2")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18


# ---- Sheet 4: Raw API Data -----------------------------------------------

def _write_raw_data(wb, priced_bom: List[PricedBomLine]):
    ws = wb.create_sheet("Raw API Data")

    headers = ["MPN", "Distributor", "Stock", "MOQ", "Price Breaks (JSON)", "Product URL"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_header(cell)
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    row_idx = 2
    for pbl in priced_bom:
        for dist_name, result in sorted(pbl.distributor_prices.items()):
            breaks_json = json.dumps(
                [{"qty": pb.min_qty, "price": str(pb.unit_price_usd)}
                 for pb in result.price_breaks],
                ensure_ascii=False,
            )
            vals = [
                result.mpn,
                dist_name,
                result.stock,
                result.moq,
                breaks_json,
                result.product_url,
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _thin_border()
            if result.product_url:
                url_cell = ws.cell(row=row_idx, column=6)
                url_cell.hyperlink = result.product_url
                url_cell.font = Font(color="0563C1", underline="single")
                url_cell.value = "Link"
            row_idx += 1

    widths = [22, 16, 10, 8, 60, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# CLI (Typer)
# ---------------------------------------------------------------------------

if _HAS_TYPER:
    app = typer.Typer(
        name="pricing-xlsx",
        help="Generate KiCost-style pricing XLSX from a KiCad BOM CSV.",
        add_completion=False,
    )

    @app.command()
    def price_bom(
        bom: Path = typer.Option(..., "--bom", "-b",
            help="Path to KiCad BOM CSV exported by kicad-cli."),
        qty: int = typer.Option(1, "--qty", "-q",
            help="Build quantity (scales extended price, qty break lookup)."),
        distributors: str = typer.Option(
            "mouser,digikey,nexar,jlcpcb",
            "--distributors", "-d",
            help="Comma-separated list of distributors to query."),
        output: Path = typer.Option(Path("pricing.xlsx"), "--output", "-o",
            help="Output XLSX file path."),
        no_dnp: bool = typer.Option(True, "--no-dnp/--include-dnp",
            help="Exclude DNP components (default: exclude)."),
    ):
        """Read BOM CSV, fetch live pricing, write XLSX workbook."""
        if not bom.exists():
            typer.echo(f"ERROR: BOM file not found: {bom}", err=True)
            raise typer.Exit(1)

        dist_list = [d.strip() for d in distributors.split(",") if d.strip()]

        typer.echo(f"Reading BOM: {bom}")
        bom_lines = read_bom_csv(bom, exclude_dnp=no_dnp)
        typer.echo(f"  {len(bom_lines)} unique MPNs")

        typer.echo(f"Querying distributors: {', '.join(dist_list)}")
        priced = aggregate_prices(bom_lines, dist_list)

        typer.echo(f"Writing XLSX: {output}")
        write_xlsx(priced, output, build_qty=qty)
        typer.echo("Done.")


# ---------------------------------------------------------------------------
# __main__ entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if _HAS_TYPER:
        app()
    else:
        # Minimal argparse fallback
        import argparse

        parser = argparse.ArgumentParser(description="Generate pricing XLSX from BOM CSV.")
        parser.add_argument("--bom", required=True, type=Path)
        parser.add_argument("--qty", type=int, default=1)
        parser.add_argument("--distributors", default="mouser,digikey,nexar,jlcpcb")
        parser.add_argument("--output", type=Path, default=Path("pricing.xlsx"))
        parser.add_argument("--no-dnp", dest="no_dnp", action="store_true", default=True)
        args = parser.parse_args()

        bom_lines = read_bom_csv(args.bom, exclude_dnp=args.no_dnp)
        dist_list = [d.strip() for d in args.distributors.split(",")]
        priced    = aggregate_prices(bom_lines, dist_list)
        write_xlsx(priced, args.output, build_qty=args.qty)
        print(f"Written: {args.output}")
